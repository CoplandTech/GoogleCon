import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from department_numbers import department_numbers

def create_xlsx(contact_list, xlsx_filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Лист 1'

    # Шрифт для всей таблицы
    default_font = Font(name='Franklin Gothic Book', size=12)

    # Заголовок
    sheet.merge_cells('A1:E1')
    title_cell = sheet['A1']
    title_cell.value = 'Список телефонов сотрудников ООО "БМУ ГЭМ"'
    title_cell.font = Font(name='Franklin Gothic Book', size=12, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Заголовки столбцов
    headers = ["ФИО", "Должность", "Внутр.", "Сотовый тел.", "Эл.почта"]
    sheet.append(headers)
    for col in range(1, 6):
        cell = sheet.cell(row=2, column=col)
        cell.font = Font(name='Franklin Gothic Book', size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Группировка контактов по департаментам
    departments = {}
    for contact in contact_list:
        department = contact.get("organizations", [{}])[0].get("department")
        if department not in departments:
            departments[department] = []
        departments[department].append(contact)

    # Функция для добавления департаментов и контактов
    def add_department_contacts(department, contacts):
        # Добавляем строку с названием департамента и номером телефона
        department_name = department
        if department in department_numbers:
            department_phone = ", ".join(department_numbers[department])
            department_name += f", городской тел. {department_phone}"
        
        sheet.append([department_name, "", "", "", ""])
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)
        department_cell = sheet.cell(row=sheet.max_row, column=1)
        department_cell.font = Font(name='Franklin Gothic Book', bold=True)
        department_cell.alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание по центру

        for contact in contacts:
            names = contact.get("names", [])
            if names:
                surname = names[0].get("familyName", "")
                first_name = names[0].get("givenName", "")
                middle_name = names[0].get("middleName", "")
                full_name = f"{surname} {first_name} {middle_name}"

                # Получение должности
                job_title = contact.get("organizations", [{}])[0].get("title", "")
                # Получение номеров телефонов
                phone_numbers = contact.get("phoneNumbers", [])
                internal_phone = ""
                corporate_phone = ""
                for phone in phone_numbers:
                    if phone.get("type") == "Короткий":
                        internal_phone = phone.get("value", "")
                    elif phone.get("type") == "Корпоративный":
                        corporate_phone = phone.get("value", "")
                
                # Получение электронной почты
                email = contact.get("emailAddresses", [{}])[0].get("value", "")
                
                row = [full_name, job_title, internal_phone, corporate_phone, email]
                sheet.append(row)
                for col in range(1, 6):
                    cell = sheet.cell(row=sheet.max_row, column=col)
                    cell.font = default_font
                    if col == 1 or col == 2 or col == 5:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Добавление контактов в файл Excel
    irkutsk_departments = ["АУП Иркутск", "ДПЭС", "Иркутский филиал", "Наладка"]
    bratsk_departments = ["АУП Братск", "Братская Площадка"]

    # Город Иркутск
    sheet.append(["город Иркутск", "", "", "", ""])
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)
    city_cell = sheet.cell(row=sheet.max_row, column=1)
    city_cell.font = Font(name='Franklin Gothic Book', bold=True, size=14)
    city_cell.alignment = Alignment(horizontal='center', vertical='center')

    for department in irkutsk_departments:
        if department in departments:
            contacts = departments.pop(department)
            contacts.sort(key=lambda x: (x.get("names", [{}])[0].get("familyName", ""), 
                                         x.get("names", [{}])[0].get("givenName", ""), 
                                         x.get("names", [{}])[0].get("middleName", "")))
            add_department_contacts(department, contacts)

    # Город Братск
    sheet.append(["город Братск", "", "", "", ""])
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=5)
    city_cell = sheet.cell(row=sheet.max_row, column=1)
    city_cell.font = Font(name='Franklin Gothic Book', bold=True, size=14)
    city_cell.alignment = Alignment(horizontal='center', vertical='center')

    for department in bratsk_departments:
        if department in departments:
            contacts = departments.pop(department)
            contacts.sort(key=lambda x: (x.get("names", [{}])[0].get("familyName", ""), 
                                         x.get("names", [{}])[0].get("givenName", ""), 
                                         x.get("names", [{}])[0].get("middleName", "")))
            add_department_contacts(department, contacts)

    # Остальные департаменты
    for department, contacts in departments.items():
        contacts.sort(key=lambda x: (x.get("names", [{}])[0].get("familyName", ""), 
                                     x.get("names", [{}])[0].get("givenName", ""), 
                                     x.get("names", [{}])[0].get("middleName", "")))
        add_department_contacts(department, contacts)

    # Добавление границ
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

    # Установка ширины колонок вручную
    column_widths = {
        'A': 40,  # Ширина для столбца "ФИО"
        'B': 55,  # Ширина для столбца "Должность"
        'C': 10,  # Ширина для столбца "Внутр."
        'D': 20,  # Ширина для столбца "Сотовый тел."
        'E': 35   # Ширина для столбца "Эл.почта"
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    workbook.save(xlsx_filename)